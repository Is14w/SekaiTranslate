import pandas as pd
import json
from pathlib import Path
import openpyxl
import re

def nameref_to_json(excel_file, output_dir=None, output_filename=None, sheet_name=None):
    """
    Convert a name reference table from Excel to JSON format using a column-by-column approach
    
    Args:
        excel_file (str): Path to the Excel file
        output_dir (str): Output directory, defaults to same directory as excel file
        output_filename (str): Output filename, defaults to excel filename with .json extension
        sheet_name (str): Sheet name to process, defaults to the first sheet
    """
    # Set up output path
    if output_dir is None:
        output_dir = Path(excel_file).parent
    else:
        output_dir = Path(output_dir)
        output_dir.mkdir(exist_ok=True)
    
    if output_filename is None:
        output_filename = f"{Path(excel_file).stem}_nameref.json"
        
    output_path = output_dir / output_filename
    
    # First load with openpyxl to get merged cells information
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    
    ws = wb[sheet_name]
    
    # Get merged cells information
    merged_cells = []
    for merged_range in ws.merged_cells.ranges:
        merged_cells.append({
            'min_row': merged_range.min_row,
            'max_row': merged_range.max_row,
            'min_col': merged_range.min_col,
            'max_col': merged_range.max_col
        })
    
    # Read with pandas for easier data handling
    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
    
    # Replace NaN with None for easier handling
    df = df.where(pd.notnull(df), None)
    
    # Initialize variables
    name_refs = []
    id_counter = 676
    
    num_rows, num_cols = df.shape
    
    # Identify all Tag_2 values in the first column (A)
    tag2_values = {}  # Maps row index to Tag_2 value
    callee_rows = {}  # Maps row index to callee name
    
    current_tag2 = None
    found_first_callee = False
    
    # First pass: Identify Tag_2 and callee values in column A
    for row in range(1, num_rows):
        # Check if this row in column A is a header row (only column A has content)
        if (df.iloc[row, 0] is not None and 
            all(pd.isna(df.iloc[row, col]) for col in range(1, num_cols))):
            
            if not found_first_callee:
                # First such row is a Tag_2 value
                current_tag2 = df.iloc[row, 0]
                tag2_values[row] = current_tag2
                found_first_callee = True
            else:
                # Could be either a new Tag_2 or a callee name
                # If it's all uppercase or contains a special character like /, it's likely a Tag_2
                value = str(df.iloc[row, 0])
                if (value.isupper() or '/' in value or 
                    any(tag in value for tag in ["Leo/need", "MORE MORE JUMP", "Vivid BAD SQUAD", 
                                                "Wonderlands×Showtime", "25时", "VIRTUAL SINGER"])):
                    current_tag2 = value
                    tag2_values[row] = current_tag2
                else:
                    # Otherwise it's a callee name
                    callee_rows[row] = df.iloc[row, 0]
    
    # Identify caller columns (columns that have a name in the first row)
    caller_cols = {}  # Maps column index to caller name
    tag1_values = {}  # Maps column index to Tag_1 value
    
    current_tag1 = None
    
    # Second pass: Identify Tag_1 values and caller columns
    for col in range(1, num_cols):
        # Skip empty cells in the first row
        if pd.isna(df.iloc[0, col]):
            continue
            
        # Check if this is a header-only column (only first row has content)
        if all(pd.isna(df.iloc[row, col]) for row in range(1, num_rows)):
            # If first time finding a header, it's the Tag_1
            if current_tag1 is None:
                current_tag1 = df.iloc[0, col]
                tag1_values[col] = current_tag1
            else:
                # Update the Tag_1 value
                current_tag1 = df.iloc[0, col]
                tag1_values[col] = current_tag1
        else:
            # This is a caller column
            caller_cols[col] = df.iloc[0, col]
    
    # Process each caller column
    for col in sorted(caller_cols.keys()):
        caller_name = caller_cols[col]
        
        # Find current Tag_1 value for this caller
        current_tag1 = None
        for c in range(col, 0, -1):
            if c in tag1_values:
                current_tag1 = tag1_values[c]
                break
        
        # Process each row for this caller
        for row in range(1, num_rows):
            # Skip rows that are Tag_2 or callee headers
            if row in tag2_values or row in callee_rows:
                continue
            
            # Find the callee for this row by looking at column A
            callee_name = df.iloc[row, 0]
            if pd.isna(callee_name):
                # If callee name is not directly in column A for this row,
                # look upward to find the closest callee name
                for r in range(row-1, -1, -1):
                    if r in callee_rows:
                        callee_name = callee_rows[r]
                        break
                    elif (df.iloc[r, 0] is not None and 
                          r not in tag2_values and 
                          not all(pd.isna(df.iloc[r, c]) for c in range(1, num_cols))):
                        callee_name = df.iloc[r, 0]
                        break
            
            # If still no callee name, skip this row
            if pd.isna(callee_name):
                continue
                
            # Find current Tag_2 value for this row
            current_tag2 = None
            for r in range(row, -1, -1):
                if r in tag2_values:
                    current_tag2 = tag2_values[r]
                    break
            
            # Check if this cell is part of a merged cell
            is_merged = False
            for mc in merged_cells:
                if (mc['min_row'] <= row+1 <= mc['max_row'] and 
                    mc['min_col'] <= col+1 <= mc['max_col']):
                    is_merged = True
                    break
            
            # Get the original value
            orig_value = df.iloc[row, col]
            
            # Handle empty cells - if both original and translation are empty, set both to null
            if pd.isna(orig_value):
                # Check if both original and translation are empty
                if col + 1 < num_cols and pd.isna(df.iloc[row, col+1]):
                    # Create entry with null values for original and translation
                    name_ref = {
                        "称呼者": caller_name,
                        "被称者": callee_name,
                        "原文": None,
                        "译文": None,
                        "id": id_counter
                    }
                    if current_tag1 is not None:
                        name_ref["Tag_1"] = current_tag1
                    if current_tag2 is not None:
                        name_ref["Tag_2"] = current_tag2
                    name_refs.append(name_ref)
                    id_counter += 1
                continue
            
            # Check next column for translation
            if col + 1 < num_cols and not pd.isna(df.iloc[row, col+1]):
                # Regular name reference with translation in next column
                trans_value = df.iloc[row, col+1]
                
                name_ref = {
                    "称呼者": caller_name,
                    "被称者": callee_name,
                    "原文": orig_value,
                    "译文": trans_value,
                    "id": id_counter
                }
                if current_tag1 is not None:
                    name_ref["Tag_1"] = current_tag1
                if current_tag2 is not None and current_tag2 != current_tag1:
                    name_ref["Tag_2"] = current_tag2
                name_refs.append(name_ref)
                id_counter += 1
            elif is_merged:
                # Self-reference (merged cell)
                name_ref = {
                    "称呼者": caller_name,
                    "被称者": callee_name,
                    "原文": orig_value,
                    "译文": orig_value,  # Same as original for self-references
                    "id": id_counter
                }
                if current_tag1 is not None:
                    name_ref["Tag_1"] = current_tag1
                if current_tag2 is not None and current_tag2 != current_tag1:
                    name_ref["Tag_2"] = current_tag2
                name_refs.append(name_ref)
                id_counter += 1
    
    # Create the final JSON structure
    json_data = {
        "人称表": name_refs
    }
    
    # Write the JSON file
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)
    
    print(f"Converted name reference table to JSON. Saved to {output_path}")
    print(f"Generated {len(name_refs)} name reference entries.")
    print(f"Found {len(tag1_values)} Tag_1 values and {len(tag2_values)} Tag_2 values")
    print(f"Found {len(caller_cols)} caller columns")

def main():
    excel_file = r"./PJS翻译资料.xlsx"  # Replace with your actual Excel file path
    output_dir = "./output"        # Output directory
    output_filename = "人称表_2.json"  # Output filename
    sheet_name = "人称表_2"           # Sheet name to process
    
    nameref_to_json(excel_file, output_dir, output_filename, sheet_name)

if __name__ == "__main__":
    main()